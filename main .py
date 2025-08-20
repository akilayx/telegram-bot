import os
import logging
import sqlite3
from datetime import datetime
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters
import openpyxl
from openpyxl.styles import Font, Alignment
import tempfile
import pandas as pd

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO)
logger = logging.getLogger(__name__)

# Get Telegram token from environment variables
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")

if not TELEGRAM_TOKEN:
    logger.error("TELEGRAM_TOKEN environment variable is required")
    exit(1)

# Database configuration
DB_NAME = "transactions.db"


# Initialize SQLite database
def init_database():
    """Initialize the SQLite database with required tables."""
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # Create transactions table
    c.execute('''CREATE TABLE IF NOT EXISTS transactions
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER NOT NULL,
                  username TEXT,
                  date TEXT NOT NULL,
                  amount REAL NOT NULL,
                  category TEXT NOT NULL,
                  description TEXT)''')

    # Create user preferences table
    c.execute('''CREATE TABLE IF NOT EXISTS user_preferences
                 (user_id INTEGER PRIMARY KEY,
                  language TEXT DEFAULT 'en')''')

    conn.commit()
    conn.close()


# Language support
LANGUAGES = {
    "en": {
        "start":
        "🤖 *Personal Finance Tracker Bot*\n\nWelcome! I help you track your income and expenses.\n\n*Available Commands:*\n📈 `/add 1000 salary` - Add income (positive amount)\n📉 `/add -250 groceries` - Add expense (negative amount)\n💰 `/balance` - Show current balance\n📊 `/export` - Export transactions to Excel\n📋 `/history` - Show recent transactions\n📈 `/report 2025-08-01 2025-08-19 [category]` - Generate date range report\n🗑️ `/clear` - Clear all transactions\n🌐 `/setlang ru` - Change language\n📂 `/categories` - Show available categories\n❓ `/help` - Show this help message\n\nStart tracking your finances now! 💼",
        "added":
        "✅ *Transaction Added*\n\n💰 Amount: `{amount:,.2f}`\n📂 Category: {category}\n📝 Description: {description}\n💰 New Balance: `{balance:,.2f}`",
        "error":
        "⚠️ Please provide an amount!\nExample: `/add 500 salary` or `/add -200 groceries`",
        "balance":
        "📊 *Your Financial Summary*\n\n💰 **Current Balance:** `{balance:,.2f}`\n📈 **Total Income:** `{income:,.2f}`\n📉 **Total Expenses:** `{expenses:,.2f}`\n📝 **Transactions:** {count}\n\n{status}",
        "export":
        "📊 Generating your financial report...",
        "report":
        "📊 *Financial Report*\n📅 Period: {start} to {end}\n📂 Category: {category}\n\n📈 **Income:** `{income:,.2f}`\n📉 **Expenses:** `{expenses:,.2f}`\n💰 **Net Total:** `{total:,.2f}`\n📝 **Transactions:** {count}",
        "lang_set":
        "✅ Language set to English",
        "categories":
        "📂 *Available Categories:*\nsalary, food, transport, entertainment, shopping, utilities, healthcare, education, other",
        "no_transactions":
        "📊 No transactions found for the specified period.",
        "cleared":
        "🗑️ *Transactions Cleared*\n\nSuccessfully removed {count} transactions.\nYour balance has been reset to 0.",
        "invalid_date":
        "❌ Invalid date format. Please use YYYY-MM-DD format.",
        "file_processed":
        "📊 File processed successfully! Added {rows} transactions.\nNew balance: {balance}",
        "file_error":
        "❌ Error processing file. Please make sure it's a valid Excel/CSV file with Date, Amount, and Category columns.",
        "help":
        "❓ *Help - Personal Finance Tracker*\n\n*Available Commands:*\n\n📈 `/add <amount> [category] [description]`\n   Add income (positive) or expense (negative)\n   Examples:\n   • `/add 1500 salary Monthly salary`\n   • `/add -75.50 food Groceries and coffee`\n\n💰 `/balance` - Show your current balance and summary\n📋 `/history` - Show your 10 most recent transactions\n📊 `/export` - Export all transactions to Excel file\n📈 `/report <start_date> <end_date> [category]` - Generate report for date range\n🗑️ `/clear` - Clear all your transactions (irreversible!)\n🌐 `/setlang <language>` - Change language (en, ru, kg)\n📂 `/categories` - Show available categories\n❓ `/help` - Show this help message\n\n*Tips:* Use positive numbers for income, negative for expenses. Descriptions are optional but helpful."
    },
    "ru": {
        "start":
        "🤖 *Бот учёта финансов*\n\nДобро пожаловать! Я помогаю отслеживать доходы и расходы.\n\n*Доступные команды:*\n📈 `/add 1000 зарплата` - Добавить доход (положительная сумма)\n📉 `/add -250 продукты` - Добавить расход (отрицательная сумма)\n💰 `/balance` - Показать текущий баланс\n📊 `/export` - Экспорт в Excel\n📋 `/history` - Показать последние транзакции\n📈 `/report 2025-08-01 2025-08-19 [категория]` - Отчёт за период\n🗑️ `/clear` - Очистить все транзакции\n🌐 `/setlang en` - Сменить язык\n📂 `/categories` - Показать категории\n❓ `/help` - Показать помощь\n\nНачните отслеживать финансы прямо сейчас! 💼",
        "added":
        "✅ *Транзакция добавлена*\n\n💰 Сумма: `{amount:,.2f}`\n📂 Категория: {category}\n📝 Описание: {description}\n💰 Новый баланс: `{balance:,.2f}`",
        "error":
        "⚠️ Укажите сумму!\nПример: `/add 500 зарплата` или `/add -200 продукты`",
        "balance":
        "📊 *Финансовая сводка*\n\n💰 **Текущий баланс:** `{balance:,.2f}`\n📈 **Общий доход:** `{income:,.2f}`\n📉 **Общие расходы:** `{expenses:,.2f}`\n📝 **Транзакций:** {count}\n\n{status}",
        "export":
        "📊 Генерация финансового отчёта...",
        "report":
        "📊 *Финансовый отчёт*\n📅 Период: {start} - {end}\n📂 Категория: {category}\n\n📈 **Доходы:** `{income:,.2f}`\n📉 **Расходы:** `{expenses:,.2f}`\n💰 **Итого:** `{total:,.2f}`\n📝 **Транзакций:** {count}",
        "lang_set":
        "✅ Язык переключён на русский",
        "categories":
        "📂 *Доступные категории:*\nзарплата, еда, транспорт, развлечения, покупки, коммуналка, здоровье, образование, прочее",
        "no_transactions":
        "📊 Транзакции за указанный период не найдены.",
        "cleared":
        "🗑️ *Транзакции очищены*\n\nУспешно удалено {count} транзакций.\nВаш баланс сброшен на 0.",
        "invalid_date":
        "❌ Неверный формат даты. Используйте формат YYYY-MM-DD.",
        "file_processed":
        "📊 Файл успешно обработан! Добавлено {rows} транзакций.\nНовый баланс: {balance}",
        "file_error":
        "❌ Ошибка обработки файла. Убедитесь, что это корректный Excel/CSV файл с колонками Date, Amount и Category.",
        "help":
        "❓ *Помощь - Учёт финансов*\n\n*Доступные команды:*\n\n📈 `/add <сумма> [категория] [описание]`\n   Добавить доход (положительное) или расход (отрицательное)\n   Примеры:\n   • `/add 1500 зарплата Месячная зарплата`\n   • `/add -75.50 еда Продукты и кофе`\n\n💰 `/balance` - Показать баланс и сводку\n📋 `/history` - Показать 10 последних транзакций\n📊 `/export` - Экспорт всех транзакций в Excel\n📈 `/report <дата_начала> <дата_конца> [категория]` - Отчёт за период\n🗑️ `/clear` - Очистить все транзакции (необратимо!)\n🌐 `/setlang <язык>` - Сменить язык (en, ru, kg)\n📂 `/categories` - Показать категории\n❓ `/help` - Показать помощь\n\n*Советы:* Используйте положительные числа для доходов, отрицательные для расходов."
    },
    "kg": {
        "start":
        "🤖 *Каржы эсеп боту*\n\nКош келдиңиз! Мен киреше жана чыгымдарды көзөмөлдөйм.\n\n*Жеткиликтүү буйруктар:*\n📈 `/add 1000 айлык` - Киреше кошуу (оң сан)\n📉 `/add -250 тамак` - Чыгым кошуу (терс сан)\n💰 `/balance` - Учурдагы балансты көрсөтүү\n📊 `/export` - Excel'ге экспорт\n📋 `/history` - Акыркы транзакцияларды көрсөтүү\n📈 `/report 2025-08-01 2025-08-19 [категория]` - Мезгил боюнча отчёт\n🗑️ `/clear` - Бардык транзакцияларды тазалоо\n🌐 `/setlang en` - Тилди өзгөртүү\n📂 `/categories` - Категорияларды көрсөтүү\n❓ `/help` - Жардамды көрсөтүү\n\nКаржыңызды азыр эле көзөмөлдөй баштаңыз! 💼",
        "added":
        "✅ *Транзакция кошулду*\n\n💰 Сумма: `{amount:,.2f}`\n📂 Категория: {category}\n📝 Сүрөттөмө: {description}\n💰 Жаңы баланс: `{balance:,.2f}`",
        "error":
        "⚠️ Суммасын көрсөтүңүз!\nМисалы: `/add 500 айлык` же `/add -200 тамак`",
        "balance":
        "📊 *Каржылык корутунду*\n\n💰 **Учурдагы баланс:** `{balance:,.2f}`\n📈 **Жалпы киреше:** `{income:,.2f}`\n📉 **Жалпы чыгым:** `{expenses:,.2f}`\n📝 **Транзакциялар:** {count}\n\n{status}",
        "export":
        "📊 Каржылык отчёт түзүлүүдө...",
        "report":
        "📊 *Каржылык отчёт*\n📅 Мезгили: {start} - {end}\n📂 Категория: {category}\n\n📈 **Киреше:** `{income:,.2f}`\n📉 **Чыгым:** `{expenses:,.2f}`\n💰 **Жыйынтык:** `{total:,.2f}`\n📝 **Транзакциялар:** {count}",
        "lang_set":
        "✅ Тил кыргызчага которулду",
        "categories":
        "📂 *Жеткиликтүү категориялар:*\nайлык, тамак, транспорт, көңүл ачуу, сатып алуу, коммуналдык, ден соолук, билим берүү, башка",
        "no_transactions":
        "📊 Көрсөтүлгөн мезгилде транзакциялар табылган жок.",
        "cleared":
        "🗑️ *Транзакциялар тазаланды*\n\n{count} транзакция ийгиликтүү өчүрүлдү.\nБалансыңыз 0'ге кайтарылды.",
        "invalid_date":
        "❌ Дата форматы туура эмес. YYYY-MM-DD форматын колдонуңуз.",
        "file_processed":
        "📊 Файл ийгиликтүү иштелип чыкты! {rows} транзакция кошулду.\nЖаңы баланс: {balance}",
        "file_error":
        "❌ Файлды иштетүүдө ката. Date, Amount жана Category колонкалары бар туура Excel/CSV файл болгонун текшериңиз.",
        "help":
        "❓ *Жардам - Каржы эсеби*\n\n*Жеткиликтүү буйруктар:*\n\n📈 `/add <сумма> [категория] [сүрөттөмө]`\n   Киреше (оң) же чыгым (терс) кошуу\n   Мисалдар:\n   • `/add 1500 айлык Айлык акы`\n   • `/add -75.50 тамак Азык-түлүк жана кофе`\n\n💰 `/balance` - Баланс жана корутундуну көрсөтүү\n📋 `/history` - Акыркы 10 транзакцияны көрсөтүү\n📊 `/export` - Бардык транзакцияларды Excel'ге экспорт\n📈 `/report <башталуу_күнү> <аяктоо_күнү> [категория]` - Мезгил боюнча отчёт\n🗑️ `/clear` - Бардык транзакцияларды тазалоо (кайтарылбайт!)\n🌐 `/setlang <тил>` - Тилди өзгөртүү (en, ru, kg)\n📂 `/categories` - Категорияларды көрсөтүү\n❓ `/help` - Жардамды көрсөтүү\n\n*Кеңештер:* Киреше үчүн оң сандарды, чыгым үчүн терс сандарды колдонуңуз."
    }
}

# Initialize database on startup
init_database()


def get_user_language(user_id: int) -> str:
    """Get user's preferred language from database."""
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT language FROM user_preferences WHERE user_id = ?",
              (user_id, ))
    result = c.fetchone()
    conn.close()
    return result[0] if result else "en"


def set_user_language(user_id: int, language: str):
    """Set user's preferred language in database."""
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        "INSERT OR REPLACE INTO user_preferences (user_id, language) VALUES (?, ?)",
        (user_id, language))
    conn.commit()
    conn.close()


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /start is issued."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)
    await update.message.reply_text(LANGUAGES[lang]["start"],
                                    parse_mode='Markdown')


async def add_transaction(update: Update,
                          context: ContextTypes.DEFAULT_TYPE) -> None:
    """Add a new transaction to the database."""
    user_id = update.effective_user.id
    username = update.effective_user.username or update.effective_user.first_name
    lang = get_user_language(user_id)

    try:
        if not context.args:
            await update.message.reply_text(LANGUAGES[lang]["error"],
                                            parse_mode='Markdown')
            return

        # Parse amount
        try:
            amount = float(context.args[0])
        except ValueError:
            await update.message.reply_text(LANGUAGES[lang]["error"],
                                            parse_mode='Markdown')
            return

        # Get category and description
        if len(context.args) > 1:
            # If second argument looks like a category, use it
            category = context.args[1].lower()
            description = " ".join(context.args[2:]) if len(
                context.args) > 2 else category
        else:
            category = "other"
            description = "No description"

        # Store transaction in database
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """INSERT INTO transactions (user_id, username, date, amount, category, description) 
                     VALUES (?, ?, ?, ?, ?, ?)""",
            (user_id, username, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             amount, category, description))
        conn.commit()

        # Calculate new balance
        c.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ?",
                  (user_id, ))
        balance = c.fetchone()[0] or 0
        conn.close()

        # Send response
        await update.message.reply_text(LANGUAGES[lang]["added"].format(
            amount=amount,
            category=category,
            description=description,
            balance=balance),
                                        parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error in add_transaction: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def show_balance(update: Update,
                       context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show current balance for the user."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        # Get all transactions for user
        c.execute("SELECT amount FROM transactions WHERE user_id = ?",
                  (user_id, ))
        amounts = [row[0] for row in c.fetchall()]

        if not amounts:
            await update.message.reply_text(
                "📊 No transactions recorded yet.\nUse `/add` to start tracking your finances!",
                parse_mode='Markdown')
            return

        total_balance = sum(amounts)
        total_income = sum(amount for amount in amounts if amount > 0)
        total_expenses = sum(amount for amount in amounts if amount < 0)
        transaction_count = len(amounts)

        status = "🟢 You're in the green!" if total_balance >= 0 else "🔴 You're in the red!"

        await update.message.reply_text(LANGUAGES[lang]["balance"].format(
            balance=total_balance,
            income=total_income,
            expenses=abs(total_expenses),
            count=transaction_count,
            status=status),
                                        parse_mode='Markdown')

        conn.close()

    except Exception as e:
        logger.error(f"Error in show_balance: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def generate_report(update: Update,
                          context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generate a financial report for a specific date range and optional category."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    if len(context.args) < 2:
        await update.message.reply_text(
            "⚠️ Format: `/report 2025-08-01 2025-08-19 [category]`",
            parse_mode='Markdown')
        return

    start_date = context.args[0]
    end_date = context.args[1]
    category = context.args[2].lower() if len(context.args) > 2 else None

    try:
        # Validate date format
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        await update.message.reply_text(LANGUAGES[lang]["invalid_date"],
                                        parse_mode='Markdown')
        return

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        if category:
            c.execute(
                """SELECT amount FROM transactions 
                        WHERE user_id = ? AND date BETWEEN ? AND ? AND category = ?""",
                (user_id, start_date + " 00:00:00", end_date + " 23:59:59",
                 category))
        else:
            c.execute(
                """SELECT amount FROM transactions 
                        WHERE user_id = ? AND date BETWEEN ? AND ?""",
                (user_id, start_date + " 00:00:00", end_date + " 23:59:59"))

        amounts = [row[0] for row in c.fetchall()]
        conn.close()

        if not amounts:
            await update.message.reply_text(LANGUAGES[lang]["no_transactions"],
                                            parse_mode='Markdown')
            return

        income = sum(amount for amount in amounts if amount > 0)
        expenses = sum(amount for amount in amounts if amount < 0)
        total = income + expenses
        count = len(amounts)

        await update.message.reply_text(LANGUAGES[lang]["report"].format(
            start=start_date,
            end=end_date,
            category=category or "all",
            income=income,
            expenses=abs(expenses),
            total=total,
            count=count),
                                        parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error in generate_report: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def set_language(update: Update,
                       context: ContextTypes.DEFAULT_TYPE) -> None:
    """Set user's preferred language."""
    user_id = update.effective_user.id

    if not context.args or context.args[0] not in LANGUAGES:
        await update.message.reply_text(
            "⚠️ Available languages: en, ru, kg\nExample: `/setlang ru`",
            parse_mode='Markdown')
        return

    new_lang = context.args[0]
    set_user_language(user_id, new_lang)

    await update.message.reply_text(LANGUAGES[new_lang]["lang_set"],
                                    parse_mode='Markdown')


async def show_categories(update: Update,
                          context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show available transaction categories."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    await update.message.reply_text(LANGUAGES[lang]["categories"],
                                    parse_mode='Markdown')


async def show_history(update: Update,
                       context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show recent transaction history."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        # Get recent transactions (last 10)
        c.execute(
            """SELECT date, amount, category, description FROM transactions 
                    WHERE user_id = ? ORDER BY date DESC LIMIT 10""",
            (user_id, ))
        transactions_data = c.fetchall()
        conn.close()

        if not transactions_data:
            await update.message.reply_text(
                "📋 *Transaction History*\n\nNo transactions found.\nUse `/add` to start tracking your finances!",
                parse_mode='Markdown')
            return

        history_message = "📋 *Recent Transactions*\n\n"

        for i, (date_str, amount, category,
                description) in enumerate(transactions_data, 1):
            # Parse date string
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            formatted_date = date_obj.strftime("%m/%d %H:%M")

            emoji = "💰" if amount > 0 else "💸"
            history_message += f"{i}. {emoji} `{amount:+,.2f}` - {description}\n    📂 {category} | 📅 {formatted_date}\n\n"

        # Check if there are more transactions
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM transactions WHERE user_id = ?",
                  (user_id, ))
        total_count = c.fetchone()[0]
        conn.close()

        if total_count > 10:
            history_message += f"... and {total_count - 10} more transactions\n"
            history_message += "Use `/export` to get all transactions in Excel format."

        await update.message.reply_text(history_message, parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error in show_history: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def export_transactions(update: Update,
                              context: ContextTypes.DEFAULT_TYPE) -> None:
    """Export user transactions to Excel file."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        # Get all transactions for user
        c.execute(
            """SELECT date, amount, category, description FROM transactions 
                    WHERE user_id = ? ORDER BY date ASC""", (user_id, ))
        transactions_data = c.fetchall()
        conn.close()

        if not transactions_data:
            await update.message.reply_text(
                "📊 No transactions to export.\nAdd some transactions first using `/add`!",
                parse_mode='Markdown')
            return

        # Send initial message
        await update.message.reply_text(LANGUAGES[lang]["export"])

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Transactions"

        # Headers
        headers = [
            "Date", "Time", "Amount", "Category", "Type", "Description",
            "Running Balance"
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        running_balance = 0
        for date_str, amount, category, description in transactions_data:
            # Parse date string
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            transaction_type = "Income" if amount > 0 else "Expense"
            running_balance += amount

            row_data = [
                date_obj.strftime("%Y-%m-%d"),
                date_obj.strftime("%H:%M:%S"), amount, category,
                transaction_type, description, running_balance
            ]
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary at the end
        last_row = ws.max_row + 2
        ws[f'A{last_row}'] = "SUMMARY"
        ws[f'A{last_row}'].font = Font(bold=True)

        total_income = sum(amount for _, amount, _, _ in transactions_data
                           if amount > 0)
        total_expenses = sum(amount for _, amount, _, _ in transactions_data
                             if amount < 0)
        final_balance = sum(amount for _, amount, _, _ in transactions_data)

        ws[f'A{last_row + 1}'] = "Total Income:"
        ws[f'B{last_row + 1}'] = total_income
        ws[f'A{last_row + 2}'] = "Total Expenses:"
        ws[f'B{last_row + 2}'] = abs(total_expenses)
        ws[f'A{last_row + 3}'] = "Final Balance:"
        ws[f'B{last_row + 3}'] = final_balance

        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False,
                                         suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)

            # Generate filename with current date
            filename = f"financial_transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            with open(tmp_file.name, 'rb') as file:
                await update.message.reply_document(
                    document=file,
                    filename=filename,
                    caption=f"📊 *Financial Report Exported*\n\n"
                    f"📝 Total Transactions: {len(transactions_data)}\n"
                    f"💰 Final Balance: `{final_balance:,.2f}`\n"
                    f"📅 Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    parse_mode='Markdown')

        # Clean up temporary file
        os.unlink(tmp_file.name)

    except Exception as e:
        logger.error(f"Error in export_transactions: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def clear_transactions(update: Update,
                             context: ContextTypes.DEFAULT_TYPE) -> None:
    """Clear all transactions for the user."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        # Get count of user transactions
        c.execute("SELECT COUNT(*) FROM transactions WHERE user_id = ?",
                  (user_id, ))
        user_transaction_count = c.fetchone()[0]

        if user_transaction_count == 0:
            await update.message.reply_text("🗑️ No transactions to clear!",
                                            parse_mode='Markdown')
            conn.close()
            return

        # Remove user's transactions
        c.execute("DELETE FROM transactions WHERE user_id = ?", (user_id, ))
        conn.commit()
        conn.close()

        await update.message.reply_text(
            LANGUAGES[lang]["cleared"].format(count=user_transaction_count),
            parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error in clear_transactions: {e}")
        await update.message.reply_text(LANGUAGES[lang]["error"])


async def help_command(update: Update,
                       context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show help information."""
    user_id = update.effective_user.id
    lang = get_user_language(user_id)
    await update.message.reply_text(LANGUAGES[lang]["help"],
                                    parse_mode='Markdown')


async def handle_file(update: Update,
                      context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle uploaded Excel/CSV files for bulk transaction import."""
    user_id = update.effective_user.id
    username = update.effective_user.username or update.effective_user.first_name
    lang = get_user_language(user_id)

    try:
        # Get file from message
        file = await update.message.document.get_file()

        # Download file to temporary location
        with tempfile.NamedTemporaryFile(delete=False,
                                         suffix='.xlsx') as tmp_file:
            await file.download_to_drive(tmp_file.name)

            # Read file with pandas
            try:
                if update.message.document.file_name.endswith('.csv'):
                    df = pd.read_csv(tmp_file.name)
                else:
                    df = pd.read_excel(tmp_file.name)
            except Exception:
                await update.message.reply_text(LANGUAGES[lang].get(
                    "file_error", "Error processing file"))
                os.unlink(tmp_file.name)
                return

            # Clean up temp file
            os.unlink(tmp_file.name)

        # Validate required columns
        required_columns = ['Date', 'Amount']
        if not all(col in df.columns for col in required_columns):
            await update.message.reply_text(LANGUAGES[lang].get(
                "file_error", "Error processing file"))
            return

        # Process each row
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        rows_added = 0
        for _, row in df.iterrows():
            try:
                # Parse date
                date_val = row.get('Date')
                if pd.isna(date_val):
                    continue

                date_str = pd.to_datetime(date_val).strftime(
                    "%Y-%m-%d %H:%M:%S")
                amount = float(row['Amount'])

                category_val = row.get('Category', 'other')
                category = str(category_val).lower(
                ) if not pd.isna(category_val) else 'other'

                description_val = row.get('Description', category)
                description = str(description_val) if not pd.isna(
                    description_val) else category

                # Insert transaction
                c.execute(
                    """INSERT INTO transactions (user_id, username, date, amount, category, description) 
                            VALUES (?, ?, ?, ?, ?, ?)""",
                    (user_id, username, date_str, amount, category,
                     description))
                rows_added += 1

            except (ValueError, TypeError):
                continue  # Skip invalid rows

        conn.commit()

        # Calculate new balance
        c.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ?",
                  (user_id, ))
        balance = c.fetchone()[0] or 0
        conn.close()

        # Send success message
        await update.message.reply_text(LANGUAGES[lang].get(
            "file_processed",
            "File processed successfully! Added {rows} transactions.\nNew balance: {balance}"
        ).format(rows=rows_added, balance=balance),
                                        parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error in handle_file: {e}")
        await update.message.reply_text(LANGUAGES[lang].get(
            "file_error", "Error processing file"))


async def error_handler(update: Update,
                        context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle errors caused by Updates."""
    logger.warning(f'Update {update} caused error {context.error}')

    if update and update.message:
        await update.message.reply_text(
            "❌ An unexpected error occurred. Please try again or contact support if the problem persists."
        )


def main() -> None:
    """Start the bot."""
    # Create the Application
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    # Register command handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add", add_transaction))
    application.add_handler(CommandHandler("balance", show_balance))
    application.add_handler(CommandHandler("history", show_history))
    application.add_handler(CommandHandler("export", export_transactions))
    application.add_handler(CommandHandler("report", generate_report))
    application.add_handler(CommandHandler("setlang", set_language))
    application.add_handler(CommandHandler("categories", show_categories))
    application.add_handler(CommandHandler("clear", clear_transactions))
    application.add_handler(CommandHandler("help", help_command))

    # Register file handler for Excel/CSV uploads
    application.add_handler(
        MessageHandler(
            filters.Document.FileExtension("xlsx")
            | filters.Document.FileExtension("xls")
            | filters.Document.FileExtension("csv"), handle_file))

    # Register error handler
    application.add_error_handler(error_handler)

    # Start the bot
    logger.info("✅ Personal Finance Tracker Bot is starting...")
    print("✅ Personal Finance Tracker Bot is running!")
    print("🔑 Bot is ready to receive commands.")
    print("📊 Financial tracking system is active.")

    # Run the bot until the user presses Ctrl-C
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
from flask import Flask
from threading import Thread

app = Flask('')

@app.route('/')
def home():
    return "Bot is running!"

def run():
    app.run(host='0.0.0.0', port=8080)

Thread(target=run).start()
